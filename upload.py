import os
import time
import pandas as pd # type: ignore
import requests # type: ignore
import re
import tempfile
from datetime import datetime
from shutil import copyfile
from openpyxl import load_workbook # type: ignore
from requests_toolbelt.multipart.encoder import MultipartEncoder # type: ignore
from bs4 import BeautifulSoup # type: ignore

# ====== KONFIGURASI ======
EXCEL_FILE = "data_siswa.xlsx"
CONFIG_FILE = "config.txt"
LOG_FILE = "log_upload.txt"
FOTO_FOLDER = "kartu_keluarga"
API_URL = "https://api-emis.kemenag.go.id/v1/students/ppdb"
DELAY = 3

# ====== UTILITAS ======
def safe_str(val):
    return str(val).strip() if pd.notna(val) else ""
def load_config(config_path="config.txt"):
    config = {}
    if os.path.exists(config_path):
        with open(config_path, "r", encoding="utf-8") as f:
            for line in f:
                if "=" in line:
                    key, value = line.strip().split("=", 1)
                    config[key.strip().lower()] = value.strip()
    return config

def backup_excel(source_path):
    backup_name = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    if os.path.exists(source_path):
        copyfile(source_path, backup_name)
        print(f"📁 Backup dibuat: {backup_name}")
        return backup_name
    else:
        raise FileNotFoundError("❌ File Excel tidak ditemukan.")

def tulis_log(message, log_path=LOG_FILE):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"{timestamp} {message}"
    print(line)
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(line + "\n")

def update_status_excel(ws, nik, col_nik_index, col_status_index, col_timestamp_index):
    for row_excel in ws.iter_rows(min_row=2, values_only=False):
        nik_excel = safe_str(row_excel[col_nik_index - 1].value)
        status_excel = safe_str(row_excel[col_status_index - 1].value).lower()

        if nik_excel == nik and status_excel in ["", "belum"]:
            row_excel[col_status_index - 1].value = "sudah"
            row_excel[col_timestamp_index - 1].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            return True
    return False

def upload_data(files, headers, url, max_retry=3):
    for attempt in range(max_retry):
        try:
            m = MultipartEncoder(fields=files)
            headers['Content-Type'] = m.content_type
            response = requests.post(url, headers=headers, data=m)
            return response
        except Exception as e:
            print(f"⚠️ Gagal upload (percobaan {attempt+1}): {e}")
            time.sleep(2)
    return None

def validasi_usia_masuk(birth_date_str, admission_date_str):
    try:
        birth_date = pd.to_datetime(birth_date_str, errors='coerce')
        admission_date = pd.to_datetime(admission_date_str, errors='coerce')
        if pd.isna(birth_date) or pd.isna(admission_date):
            return False
        return birth_date <= admission_date
    except:
        return False
    
#====== Format Nomor HP ======
def format_nomor_hp(nomor):
    nomor = safe_str(nomor)
    if not nomor:
        return "", "false"

    # Hilangkan karakter non-digit
    nomor = re.sub(r"[^\d]", "", nomor)

    # Jika sudah diawali dengan 62 dan panjang cukup
    if re.match(r"^62\d{8,}$", nomor):
        return nomor, "true"

    # Jika diawali dengan 08 dan panjang cukup
    elif re.match(r"^08\d{8,}$", nomor):
        nomor_formatted = "62" + nomor[1:]
        return nomor_formatted, "true"

    # Format tidak valid
    return "", "false"

def format_tanggal(tanggal_str, min_year=1950, max_year=None):
    try:
        tanggal = pd.to_datetime(tanggal_str, errors='coerce')
        if pd.isna(tanggal):
            return "", False
        tahun = tanggal.year
        batas_atas = max_year if max_year else datetime.now().year
        if tahun < min_year or tahun > batas_atas:
            return "", False
        return tanggal.strftime("%Y-%m-%d"), True
    except:
        return "", False

def download_file_from_url(url, full_name):
    try:
        response = requests.get(url, timeout=30)
        if response.status_code != 200:
            print(f"❌ Gagal download file dari URL: {url} (Status: {response.status_code})")
            return None, None
        
        content_type = response.headers.get('Content-Type', '').lower()
        
        if 'pdf' in content_type or url.lower().endswith('.pdf'):
            extension = 'pdf'
            mime_type = 'application/pdf'
        elif 'image' in content_type or any(url.lower().endswith(ext) for ext in ['.jpg', '.jpeg', '.png']):
            if 'png' in content_type or url.lower().endswith('.png'):
                extension = 'png'
                mime_type = 'image/png'
            else:
                extension = 'jpg'
                mime_type = 'image/jpeg'
        else:
            extension = 'jpg'
            mime_type = 'image/jpeg'
        
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=f'.{extension}')
        temp_file.write(response.content)
        temp_file.close()
        
        return temp_file.name, mime_type
    except Exception as e:
        print(f"❌ Error saat download file dari URL: {e}")
        return None, None

def get_kk_file(full_name, kk_url_or_empty):
    kk_url = safe_str(kk_url_or_empty)
    
    if kk_url and (kk_url.startswith('http://') or kk_url.startswith('https://')):
        print(f"📥 Download kartu keluarga dari URL: {kk_url}")
        temp_path, mime_type = download_file_from_url(kk_url, full_name)
        if temp_path:
            try:
                foto_file = open(temp_path, "rb")
                return foto_file, temp_path, mime_type, True
            except Exception as e:
                print(f"❌ Gagal membuka file yang didownload: {e}")
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                return None, None, None, False
        else:
            return None, None, None, False
    else:
        foto_filename = os.path.join(FOTO_FOLDER, f"{full_name}.jpg")
        if not os.path.exists(foto_filename):
            print(f"❌ File foto {foto_filename} tidak ditemukan di folder lokal.")
            return None, None, None, False
        try:
            foto_file = open(foto_filename, "rb")
            return foto_file, foto_filename, 'image/jpeg', False
        except Exception as e:
            print(f"❌ Gagal membuka file foto: {e}")
            return None, None, None, False

# ====== LOAD KONFIGURASI ======
config = load_config(CONFIG_FILE)
token = config.get("token", "")
academic_year_id = config.get("academic_year_id", "18")
admission_date = config.get("admission_date", "2025-07-14")
    
# ====== FUNGSI BUILD FILES (tidak diubah) ======
def build_files(data):
    """Membangun payload files sesuai format"""
    full_name = str(data['full_name']).strip()

    gender_raw = str(data['gender']).strip().lower()
    gender_id = "2" if gender_raw in ["perempuan", "p", "wanita", "female", "2"] else "1"

    father_phone_number_raw = data['father_phone_number']
    father_phone_number, father_have_phone = format_nomor_hp(father_phone_number_raw)

    mother_phone_number_raw = data['mother_phone_number']
    mother_phone_number, mother_have_phone = format_nomor_hp(mother_phone_number_raw)

    birth_date, valid_birth = format_tanggal(data.get("birth_date"),min_year=2005)
    father_birth_date, valid_father = format_tanggal(data.get("father_birth_date"),min_year=1950)
    mother_birth_date, valid_mother = format_tanggal(data.get("mother_birth_date"),min_year=1950)

    if not validasi_usia_masuk(birth_date, admission_date):
        print(f"❌ Tanggal lahir siswa melebihi tanggal masuk: {full_name} | birth_date: {birth_date} > admission_date: {admission_date}")
        tulis_log(f"❌ Tanggal lahir siswa melebihi tanggal masuk: {full_name} | birth_date: {birth_date} > admission_date: {admission_date}")
        return None

    kk_url = data.get('kartu_keluarga', '')
    foto_file, foto_filename, mime_type, is_temp = get_kk_file(full_name, kk_url)
    
    if not foto_file:
        print(f"❌ Kartu keluarga tidak ditemukan (cek kolom kartu_keluarga atau folder {FOTO_FOLDER})")
        return None

    if not full_name or full_name.lower() == 'nan':
        print(f"❌ Nama siswa tidak valid, batalkan input.")
        return None
    if not valid_birth:
        print(f"❌ Tanggal lahir siswa tidak valid: {full_name} | birth_date: {data.get('birth_date')}")
        return None

    if not valid_father:
        print(f"❌ Tanggal lahir ayah tidak valid: {full_name} | father_birth_date: {data.get('father_birth_date')}")
        return None

    if not valid_mother:
        print(f"❌ Tanggal lahir ibu tidak valid: {full_name} | mother_birth_date: {data.get('mother_birth_date')}")
        return None


    files = {
        'full_name': (None, data['full_name']),
        'nationality': (None, 'wni'),
        'have_nik': (None, 'true'),
        'nik': (None, str(data['nik'])),
        'have_nisn': (None, 'false'),
        'm_gender_id': (None, gender_id),
        'birth_place': (None, data['birth_place']),
        'birth_date': (None, birth_date),
        'siblings_num': (None, str(data['siblings_num'])),
        'child_of_num': (None, str(data['child_of_num'])),
        'm_residence_status_id': (None, str(data['m_residence_status_id'])),
        'm_province_id': (None, str(data['m_province_id'])),
        'm_city_id': (None, str(data['m_city_id'])),
        'm_district_id': (None, str(data['m_district_id'])),
        'm_subdistrict_id': (None, str(data['m_subdistrict_id'])),
        'address': (None, data['address']),
        'postal_code_num': (None, str(data['postal_code_num'])),
        'have_handphone': (None, 'false'),
        'm_level_id': (None, str(data['m_level_id'])),
        'kk_num': (None, str(data['kk_num'])),
        'family_head_name': (None, data['family_head_name']),
        'upload_kk': (foto_filename, foto_file, mime_type),


        # father
        'father_full_name': (None, data['father_full_name']),
        'father_m_life_status_id': (None, str(data['father_m_life_status_id'])),
        'father_nationality': (None, 'wni'),
        'father_nik': (None, str(data['father_nik'])),
        'father_birth_place': (None, data['father_birth_place']),
        'father_birth_date': (None, father_birth_date),
        'father_m_last_education_id': (None, str(data['father_m_last_education_id'])),
        'father_m_job_id': (None, str(data['father_m_job_id'])),
        'father_m_average_income_per_month_id': (None, str(data['father_m_average_income_per_month_id'])),
        'father_domicile': (None, 'Dalam Negeri'),
        'father_m_residence_status_id': (None, str(data['m_residence_status_id'])),
        'father_m_province_id': (None, str(data['m_province_id'])),
        'father_m_city_id': (None, str(data['m_city_id'])),
        'father_m_district_id': (None, str(data['m_district_id'])),
        'father_m_sub_district_id': (None, str(data['m_subdistrict_id'])),
        'father_address': (None, data['address']),
        'father_postal_code': (None, str(data['postal_code_num'])),
        'father_phone_number': (None, father_phone_number),
        'father_have_phone_number': (None,father_have_phone),
        'father_kk_file': (foto_filename, foto_file, mime_type),

        # mother
        'mother_residence': (None, '1'),
        'mother_full_name': (None, data['mother_full_name']),
        'mother_m_life_status_id': (None, str(data['mother_m_life_status_id'])),
        'mother_nationality': (None, 'wni'),
        'mother_nik': (None, str(data['mother_nik'])),
        'mother_birth_place': (None, data['mother_birth_place']),
        'mother_birth_date': (None, mother_birth_date),
        'mother_m_last_education_id': (None, str(data['mother_m_last_education_id'])),
        'mother_m_job_id': (None, str(data['mother_m_job_id'])),
        'mother_m_average_income_per_month_id': (None, str(data['mother_m_average_income_per_month_id'])),
        'mother_domicile': (None, 'Dalam Negeri'),
        'mother_m_residence_status_id': (None, str(data['m_residence_status_id'])),
        'mother_m_province_id': (None, str(data['m_province_id'])),
        'mother_m_city_id': (None, str(data['m_city_id'])),
        'mother_m_district_id': (None, str(data['m_district_id'])),
        'mother_m_sub_district_id': (None, str(data['m_subdistrict_id'])),
        'mother_address': (None, data['address']),
        'mother_postal_code': (None, str(data['postal_code_num'])),
        'mother_phone_number': (None,mother_phone_number),
        'mother_have_phone_number': (None, mother_have_phone),
        'mother_kk_file': (foto_filename, foto_file, mime_type),

        # wali
        'wali': (None, 'Sama dengan ayah kandung'),
        'wali_full_name': (None, data['father_full_name']),
        'wali_m_life_status_id': (None, str(data['father_m_life_status_id'])),
        'wali_nationality': (None, 'wni'),
        'wali_nik': (None, str(data['father_nik'])),
        'wali_birth_place': (None, data['father_birth_place']),
        'wali_birth_date': (None, str(data['father_birth_date'])),
        'wali_m_last_education_id': (None, str(data['father_m_last_education_id'])),
        'wali_m_job_id': (None, str(data['father_m_job_id'])),
        'wali_m_average_income_per_month_id': (None, str(data['father_m_average_income_per_month_id'])),
        'wali_domicile': (None, 'Dalam Negeri'),
        'wali_m_residence_status_id': (None, str(data['m_residence_status_id'])),
        'wali_m_province_id': (None, str(data['m_province_id'])),
        'wali_m_city_id': (None, str(data['m_city_id'])),
        'wali_m_district_id': (None, str(data['m_district_id'])),
        'wali_m_sub_district_id': (None, str(data['m_subdistrict_id'])),
        'wali_address': (None, data['address']),
        'wali_postal_code': (None, str(data['postal_code_num'])),
        'wali_phone_number': (None, ''),   # kosong
        'wali_have_phone_number': (None, 'false'),
        'wali_kk_file': (foto_filename, foto_file, mime_type),

        # umum
        'm_religion_id': (None, '1'), # Agama Islam
        'm_fund_source_id': (None, '1'), # Biaya Sekolah Dibayai Orang Tua
        'm_special_need_id': (None, '1'), # Kebutuhan Khusus Biasa
        'm_disability_id': (None, '1'), # Disabilitas Tidak
        'm_transportation_id': (None, '3'), # Sepeda Motor
        'm_residence_distance_id': (None, '1'), # Jarak 1 - 5 Km
        'm_interval_time_id': (None, '1'), # Waktu Tempuh 1 - 10 Menit
        'admission_date': (None, admission_date), # Tanggal Masuk Sekolah
        'academic_year_id': (None, academic_year_id), # ID Tahun Ajaran
        'is_pontren': (None, '0'),
    }
    return files, foto_file, foto_filename, is_temp

def tampilkan_ringkasan(data):
    print("\n📋 Ringkasan Data Siswa:")
    print(f"Nama Lengkap        : {data['full_name']}")
    print(f"NIK                 : {data['nik']}")
    print(f"Tempat, Tgl Lahir   : {data['birth_place']}, {data['birth_date']}")
    print(f"Alamat              : {data['address']}")
    print(f"Jenis Kelamin       : {data['gender']}")
    print(f"Nama Ayah           : {data['father_full_name']}")
    print(f"Nama Ibu            : {data['mother_full_name']}")
    print(f"File Kartu Keluarga : {str(data['full_name']).strip()}.jpg")

def safe_save_workbook(wb, path, max_retry=3):
    for attempt in range(max_retry):
        try:
            wb.save(path)
            return True
        except Exception as e:
            print(f"⚠️ Gagal menyimpan Excel (percobaan {attempt+1}): {e}")
            time.sleep(2)
    print("❌ Gagal menyimpan Excel setelah beberapa percobaan.")
    return False

def format_kode_wilayah_for_postal(kode):
    """Format kode wilayah untuk pencarian kode pos"""
    kode = str(kode).strip().replace(".", "")
    if len(kode) >= 12:
        return f"{kode[0:2]}.{kode[2:4]}.{kode[4:6]}.{kode[6:10]}"
    elif len(kode) == 10:
        return f"{kode[0:2]}.{kode[2:4]}.{kode[4:6]}.{kode[6:10]}"
    elif len(kode) == 6:
        return f"{kode[0:2]}.{kode[2:4]}.{kode[4:6]}"
    elif len(kode) == 4:
        return f"{kode[0:2]}.{kode[2:4]}"
    else:
        return kode

def cari_kodepos_web(kode_wilayah):
    """Cari kode pos dengan scraping web"""
    try:
        kode_bertitik = format_kode_wilayah_for_postal(kode_wilayah)
        url = f"https://kodepos.nomor.net/_kodepos.php?_i=cari-kodepos&jobs={kode_bertitik}"
        
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
        response = requests.get(url, headers=headers, timeout=10)
        
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            tables = soup.find_all('table')
            
            for table in tables:
                rows = table.find_all('tr')
                for row in rows:
                    cols = row.find_all(['td', 'th'])
                    
                    for i in range(len(cols) - 6):
                        col_text = cols[i].get_text(strip=True)
                        
                        if col_text.isdigit() and len(col_text) == 5:
                            try:
                                kode_wil = cols[i+2].get_text(strip=True) if i+2 < len(cols) else ''
                                if kode_wil and '.' in kode_wil:
                                    return col_text
                            except:
                                pass
        return None
    except:
        return None

def autofill_postal_codes_pre_upload():
    """Auto-fill kode pos sebelum upload"""
    print("\n" + "="*60)
    print("AUTO-FILL KODE POS")
    print("="*60)
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    postal_col_idx = headers.index('postal_code_num') + 1 if 'postal_code_num' in headers else None
    subdistrict_col_idx = headers.index('m_subdistrict_id') + 1 if 'm_subdistrict_id' in headers else None
    fullname_col_idx = headers.index('full_name') + 1 if 'full_name' in headers else None
    
    if not postal_col_idx or not subdistrict_col_idx:
        print("❌ Kolom yang diperlukan tidak ditemukan")
        print("="*60 + "\n")
        return
    
    rows_to_update = []
    for row_idx in range(2, ws.max_row + 1):
        postal_val = ws.cell(row_idx, postal_col_idx).value
        if postal_val is None or str(postal_val).strip() == '':
            rows_to_update.append(row_idx)
    
    if len(rows_to_update) == 0:
        print("✅ Semua kode pos sudah terisi")
        print("="*60 + "\n")
        wb.close()
        return
    
    print(f"📋 Ditemukan {len(rows_to_update)} baris dengan kode pos kosong")
    
    lanjut = input("➡️ Lanjutkan auto-fill kode pos? (Y/N): ").strip().lower()
    if lanjut != 'y':
        print("⏭️ Auto-fill dibatalkan, lanjut ke upload")
        print("="*60 + "\n")
        wb.close()
        return
    
    updated_count = 0
    
    for row_idx in rows_to_update:
        full_name = ws.cell(row_idx, fullname_col_idx).value if fullname_col_idx else 'N/A'
        m_subdistrict_id = ws.cell(row_idx, subdistrict_col_idx).value
        
        if not m_subdistrict_id:
            continue
        
        m_subdistrict_id_str = str(m_subdistrict_id).strip()
        
        print(f"[PROSES] {full_name} - Kode wilayah: {m_subdistrict_id_str}", end=" ")
        
        kode_pos = cari_kodepos_web(m_subdistrict_id_str)
        
        if kode_pos:
            ws.cell(row_idx, postal_col_idx).value = kode_pos
            ws.cell(row_idx, postal_col_idx).number_format = '0'
            print(f"✅ Kode pos: {kode_pos}")
            updated_count += 1
        else:
            print("❌ Tidak ditemukan")
        
        time.sleep(1)
    
    if updated_count > 0:
        wb.save(EXCEL_FILE)
        print(f"\n✅ Berhasil update {updated_count} kode pos")
        tulis_log(f"Auto-fill kode pos: {updated_count} baris berhasil diupdate")
    else:
        print("\n⚠️ Tidak ada kode pos yang berhasil diupdate")
    
    wb.close()
    print("="*60 + "\n")

def format_phone_numbers_in_excel():
    """Format nomor telepon ayah dan ibu ke format 62xxx di Excel"""
    print("\n" + "="*60)
    print("FORMAT NOMOR TELEPON")
    print("="*60)
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    father_phone_idx = headers.index('father_phone_number') + 1 if 'father_phone_number' in headers else None
    mother_phone_idx = headers.index('mother_phone_number') + 1 if 'mother_phone_number' in headers else None
    fullname_col_idx = headers.index('full_name') + 1 if 'full_name' in headers else None
    
    if not father_phone_idx and not mother_phone_idx:
        print("❌ Kolom nomor telepon tidak ditemukan")
        print("="*60 + "\n")
        wb.close()
        return
    
    updated_count = 0
    
    for row_idx in range(2, ws.max_row + 1):
        full_name = ws.cell(row_idx, fullname_col_idx).value if fullname_col_idx else 'N/A'
        
        if father_phone_idx:
            father_phone = ws.cell(row_idx, father_phone_idx).value
            if father_phone:
                nomor_str = str(father_phone).strip()
                nomor_clean = re.sub(r"[^\d]", "", nomor_str)
                
                if nomor_clean:
                    if re.match(r"^62\d{8,}$", nomor_clean):
                        formatted = nomor_clean
                    elif re.match(r"^0\d{8,}$", nomor_clean):
                        formatted = "62" + nomor_clean[1:]
                        ws.cell(row_idx, father_phone_idx).value = formatted
                        updated_count += 1
                        print(f"✅ {full_name} - Ayah: {nomor_str} -> {formatted}")
                    elif re.match(r"^8\d{8,}$", nomor_clean):
                        formatted = "62" + nomor_clean
                        ws.cell(row_idx, father_phone_idx).value = formatted
                        updated_count += 1
                        print(f"✅ {full_name} - Ayah: {nomor_str} -> {formatted}")
        
        if mother_phone_idx:
            mother_phone = ws.cell(row_idx, mother_phone_idx).value
            if mother_phone:
                nomor_str = str(mother_phone).strip()
                nomor_clean = re.sub(r"[^\d]", "", nomor_str)
                
                if nomor_clean:
                    if re.match(r"^62\d{8,}$", nomor_clean):
                        formatted = nomor_clean
                    elif re.match(r"^0\d{8,}$", nomor_clean):
                        formatted = "62" + nomor_clean[1:]
                        ws.cell(row_idx, mother_phone_idx).value = formatted
                        updated_count += 1
                        print(f"✅ {full_name} - Ibu: {nomor_str} -> {formatted}")
                    elif re.match(r"^8\d{8,}$", nomor_clean):
                        formatted = "62" + nomor_clean
                        ws.cell(row_idx, mother_phone_idx).value = formatted
                        updated_count += 1
                        print(f"✅ {full_name} - Ibu: {nomor_str} -> {formatted}")
    
    if updated_count > 0:
        wb.save(EXCEL_FILE)
        print(f"\n✅ Berhasil format {updated_count} nomor telepon")
        tulis_log(f"Format nomor telepon: {updated_count} nomor berhasil diupdate")
    else:
        print("\n✅ Semua nomor telepon sudah dalam format yang benar")
    
    wb.close()
    print("="*60 + "\n")

# ====== MAIN PROSES ======
def main():
    try:
        backup_excel(EXCEL_FILE)

    except Exception as e:
        print(e)
        return

    if not token:
        print("❌ Token tidak ditemukan di file token.txt.")
        return
    
    autofill_postal_codes_pre_upload()
    format_phone_numbers_in_excel()
    
    headers = {
        'accept': 'application/json',
        'authorization': f'Bearer {token}',
        'origin': 'https://emis.kemenag.go.id',
        'referer': 'https://emis.kemenag.go.id/',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
    }

    # ✅ Validasi format admission_date
    try:
        admission_date_parsed = pd.to_datetime(admission_date, errors='raise')
    except Exception as e:
        print(f"❌ Format admission_date tidak valid: {admission_date}")
        tulis_log(f"❌ Format admission_date tidak valid: {admission_date}")
        return
    

    df = pd.read_excel(EXCEL_FILE).fillna("")
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # ====== VALIDASI DUPLIKAT NIK ======
    duplikat_nik_mask = df['nik'].duplicated(keep=False)
    nik_duplikat_set = set(df.loc[duplikat_nik_mask, 'nik'])

    if nik_duplikat_set:
        print("⚠️ Terdapat NIK duplikat di file Excel. Baris berikut akan dilewati:")
        for nik in nik_duplikat_set:
            siswa_duplikat = df[df['nik'] == nik]
            for _, row in siswa_duplikat.iterrows():
                print(f"🔁 {row['full_name']} | NIK: {row['nik']}")
                tulis_log(f"🔁 Duplikat NIK dilewati: {row['full_name']} | NIK: {row['nik']}")

    col_nik_index = df.columns.get_loc("nik") + 1
    col_status_index = df.columns.get_loc("status") + 1
    
    if "upload_timestamp" not in [col.lower() for col in df.columns]:
        ws.cell(row=1, column=ws.max_column + 1).value = "upload_timestamp"
        col_timestamp_index = ws.max_column
    
    jumlah_sukses = 0
    jumlah_gagal = 0
    jumlah_dilewati = 0
    jumlah_dilewati_tanggal = 0
    jumlah_dilewati_nik = 0
    jumlah_dilewati_duplikat = 0

    for idx, row in df.iterrows():
        nik = safe_str(row.get("nik"))
        status = safe_str(row.get("status")).lower()

        if len(nik) != 16 or not nik.isdigit():
            tulis_log(f"❌ NIK tidak valid: {row['full_name']} | NIK: {nik}")
            jumlah_dilewati_nik += 1
            continue

        if nik in nik_duplikat_set:
            jumlah_dilewati_duplikat += 1
            continue

        if status not in ["", "belum"]:
            continue

        print(f"\n🚀 Upload siswa ke-{idx+1}: {row['full_name']} (NIK: {nik})")
        files_and_file = build_files(row)
        if files_and_file is None:
            tulis_log(f"⏭️ Dilewati karena data tidak valid: {row['full_name']}")
            jumlah_dilewati_tanggal += 1
            continue
        files, foto_file, foto_filename, is_temp = files_and_file
        tampilkan_ringkasan(row)
        lanjut = input("➡️ Lanjutkan upload? (Y/N): ").strip().lower()
        if lanjut != 'y':
            print("⏭️ Upload dibatalkan oleh pengguna.")
            if foto_file:
                foto_file.close()
            if is_temp and os.path.exists(foto_filename):
                os.remove(foto_filename)
            jumlah_dilewati += 1
            continue

        response = upload_data(files, headers, API_URL)
        full_name = row.get("full_name", "❓ Nama tidak ditemukan")

        if response and response.status_code == 200:
            tulis_log(f"✅ {full_name} | Status {response.status_code} | {response.text}")
            jumlah_sukses += 1
            if update_status_excel(ws, nik, col_nik_index, col_status_index, col_timestamp_index):
                safe_save_workbook(wb, EXCEL_FILE)
        elif response:
            tulis_log(f"⚠️ {full_name} | Status {response.status_code} | {response.text}")
            jumlah_gagal += 1
        else:
            tulis_log(f"❌ {full_name} | Gagal upload setelah beberapa percobaan.")
            jumlah_gagal += 1

        if foto_file:
            foto_file.close()
        if is_temp and os.path.exists(foto_filename):
            os.remove(foto_filename)
        time.sleep(DELAY)

    print("\n📁 Proses upload selesai. Status diperbarui langsung di file 'data_siswa.xlsx'")
    print("\n📊 Ringkasan Harian Upload:")
    print(f"✅ Berhasil upload     : {jumlah_sukses}")
    print(f"❌ Gagal upload        : {jumlah_gagal}")
    print(f"⏭️ Dilewati (manual)   : {jumlah_dilewati}")
    print(f"⏭️ Dilewati (tanggal)  : {jumlah_dilewati_tanggal}")
    print(f"⏭️ Dilewati (NIK salah): {jumlah_dilewati_nik}")
    print(f"⏭️ Dilewati (duplikat) : {jumlah_dilewati_duplikat}")
    
    tulis_log("📊 Ringkasan Harian Upload:" )
    tulis_log(f"✅ Berhasil upload     : {jumlah_sukses}")
    tulis_log(f"❌ Gagal upload        : {jumlah_gagal}")
    tulis_log(f"⏭️ Dilewati (manual)   : {jumlah_dilewati}")
    tulis_log(f"⏭️ Dilewati (tanggal)  : {jumlah_dilewati_tanggal}")
    tulis_log(f"⏭️ Dilewati (NIK salah): {jumlah_dilewati_nik}")
    tulis_log(f"⏭️ Dilewati (duplikat) : {jumlah_dilewati_duplikat}")

if __name__ == "__main__":
    main()