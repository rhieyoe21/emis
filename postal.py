import time
import requests # type: ignore
import pandas as pd # type: ignore
import re
from bs4 import BeautifulSoup # type: ignore
from openpyxl import load_workbook # type: ignore
from config import EXCEL_FILE
from utils import tulis_log, format_kode_wilayah_for_postal, parse_date_safe

def cari_kodepos_web(kode_wilayah):
    """Cari kode pos dengan scraping web"""
    try:
        kode_bertitik = format_kode_wilayah_for_postal(kode_wilayah)
        url = f"https://kodepos.nomor.net/_kodepos.php?_i=cari-kodepos&jobs={kode_bertitik}"
        
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
        response = requests.get(url, headers=headers, timeout=10)
        
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            tables = soup.find_all('table')
            
            for table in tables:
                rows = table.find_all('tr')
                for row in rows:
                    cols = row.find_all(['td', 'th'])
                    
                    for i in range(len(cols) - 6):
                        col_text = cols[i].get_text(strip=True)
                        
                        if col_text.isdigit() and len(col_text) == 5:
                            try:
                                kode_wil = cols[i+2].get_text(strip=True) if i+2 < len(cols) else ''
                                if kode_wil and '.' in kode_wil:
                                    return col_text
                            except:
                                pass
        return None
    except:
        return None

def autofill_postal_codes_pre_upload():
    """Auto-fill kode pos sebelum upload"""
    print("\n" + "="*60)
    print("AUTO-FILL KODE POS")
    print("="*60)
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    postal_col_idx = headers.index('postal_code_num') + 1 if 'postal_code_num' in headers else None
    subdistrict_col_idx = headers.index('m_subdistrict_id') + 1 if 'm_subdistrict_id' in headers else None
    fullname_col_idx = headers.index('full_name') + 1 if 'full_name' in headers else None
    
    if not postal_col_idx or not subdistrict_col_idx:
        print("❌ Kolom yang diperlukan tidak ditemukan")
        print("="*60 + "\n")
        wb.close()
        return
    
    rows_to_update = []
    for row_idx in range(2, ws.max_row + 1):
        postal_val = ws.cell(row_idx, postal_col_idx).value
        if postal_val is None or str(postal_val).strip() == '':
            rows_to_update.append(row_idx)
    
    if len(rows_to_update) == 0:
        print("✅ Semua kode pos sudah terisi")
        print("="*60 + "\n")
        wb.close()
        return
    
    print(f"📋 Ditemukan {len(rows_to_update)} baris dengan kode pos kosong")
    
    lanjut = input("➡️ Lanjutkan auto-fill kode pos? (Y/N): ").strip().lower()
    if lanjut != 'y':
        print("⏭️ Auto-fill dibatalkan, lanjut ke upload")
        print("="*60 + "\n")
        wb.close()
        return
    
    updated_count = 0
    
    for row_idx in rows_to_update:
        full_name = ws.cell(row_idx, fullname_col_idx).value if fullname_col_idx else 'N/A'
        m_subdistrict_id = ws.cell(row_idx, subdistrict_col_idx).value
        
        if not m_subdistrict_id:
            continue
        
        m_subdistrict_id_str = str(m_subdistrict_id).strip()
        
        print(f"[PROSES] {full_name} - Kode wilayah: {m_subdistrict_id_str}", end=" ")
        
        kode_pos = cari_kodepos_web(m_subdistrict_id_str)
        
        if kode_pos:
            ws.cell(row_idx, postal_col_idx).value = kode_pos
            ws.cell(row_idx, postal_col_idx).number_format = '0'
            print(f"✅ Kode pos: {kode_pos}")
            updated_count += 1
        else:
            print("❌ Tidak ditemukan")
        
        time.sleep(1)
    
    if updated_count > 0:
        wb.save(EXCEL_FILE)
        print(f"\n✅ Berhasil update {updated_count} kode pos")
        tulis_log(f"Auto-fill kode pos: {updated_count} baris berhasil diupdate")
    else:
        print("\n⚠️ Tidak ada kode pos yang berhasil diupdate")
    
    wb.close()
    print("="*60 + "\n")

def format_phone_numbers_in_excel():
    """Format nomor telepon ayah dan ibu ke format 62xxx di Excel"""
    print("\n" + "="*60)
    print("FORMAT NOMOR TELEPON")
    print("="*60)
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    father_phone_idx = headers.index('father_phone_number') + 1 if 'father_phone_number' in headers else None
    mother_phone_idx = headers.index('mother_phone_number') + 1 if 'mother_phone_number' in headers else None
    fullname_col_idx = headers.index('full_name') + 1 if 'full_name' in headers else None
    
    if not father_phone_idx and not mother_phone_idx:
        print("❌ Kolom nomor telepon tidak ditemukan")
        print("="*60 + "\n")
        wb.close()
        return
    
    updated_count = 0
    
    for row_idx in range(2, ws.max_row + 1):
        full_name = ws.cell(row_idx, fullname_col_idx).value if fullname_col_idx else 'N/A'
        
        if father_phone_idx:
            father_phone = ws.cell(row_idx, father_phone_idx).value
            if father_phone:
                nomor_str = str(father_phone).strip()
                nomor_clean = re.sub(r"[^\d]", "", nomor_str)
                
                if nomor_clean:
                    if re.match(r"^62\d{8,}$", nomor_clean):
                        formatted = nomor_clean
                    elif re.match(r"^0\d{8,}$", nomor_clean):
                        formatted = "62" + nomor_clean[1:]
                        ws.cell(row_idx, father_phone_idx).value = formatted
                        updated_count += 1
                        print(f"✅ {full_name} - Ayah: {nomor_str} -> {formatted}")
                    elif re.match(r"^8\d{8,}$", nomor_clean):
                        formatted = "62" + nomor_clean
                        ws.cell(row_idx, father_phone_idx).value = formatted
                        updated_count += 1
                        print(f"✅ {full_name} - Ayah: {nomor_str} -> {formatted}")
        
        if mother_phone_idx:
            mother_phone = ws.cell(row_idx, mother_phone_idx).value
            if mother_phone:
                nomor_str = str(mother_phone).strip()
                nomor_clean = re.sub(r"[^\d]", "", nomor_str)
                
                if nomor_clean:
                    if re.match(r"^62\d{8,}$", nomor_clean):
                        formatted = nomor_clean
                    elif re.match(r"^0\d{8,}$", nomor_clean):
                        formatted = "62" + nomor_clean[1:]
                        ws.cell(row_idx, mother_phone_idx).value = formatted
                        updated_count += 1
                        print(f"✅ {full_name} - Ibu: {nomor_str} -> {formatted}")
                    elif re.match(r"^8\d{8,}$", nomor_clean):
                        formatted = "62" + nomor_clean
                        ws.cell(row_idx, mother_phone_idx).value = formatted
                        updated_count += 1
                        print(f"✅ {full_name} - Ibu: {nomor_str} -> {formatted}")
    
    if updated_count > 0:
        wb.save(EXCEL_FILE)
        print(f"\n✅ Berhasil format {updated_count} nomor telepon")
        tulis_log(f"Format nomor telepon: {updated_count} nomor berhasil diupdate")
    else:
        print("\n✅ Semua nomor telepon sudah dalam format yang benar")
    
    wb.close()
    print("="*60 + "\n")

def format_birthdates_in_excel():
    """Format tanggal lahir siswa, ayah, dan ibu ke format date yyyy-mm-dd di Excel"""
    print("\n" + "="*60)
    print("FORMAT TANGGAL LAHIR (YYYY-MM-DD)")
    print("="*60)
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    birth_col_idx = headers.index('birth_date') + 1 if 'birth_date' in headers else None
    father_birth_col_idx = headers.index('father_birth_date') + 1 if 'father_birth_date' in headers else None
    mother_birth_col_idx = headers.index('mother_birth_date') + 1 if 'mother_birth_date' in headers else None
    fullname_col_idx = headers.index('full_name') + 1 if 'full_name' in headers else None
    
    if not birth_col_idx and not father_birth_col_idx and not mother_birth_col_idx:
        print("❌ Kolom tanggal lahir tidak ditemukan")
        print("="*60 + "\n")
        wb.close()
        return
        
    updated_count = 0
    
    for row_idx in range(2, ws.max_row + 1):
        full_name = ws.cell(row_idx, fullname_col_idx).value if fullname_col_idx else 'N/A'
        
        for label, col_idx in [("Siswa", birth_col_idx), ("Ayah", father_birth_col_idx), ("Ibu", mother_birth_col_idx)]:
            if col_idx:
                cell = ws.cell(row_idx, col_idx)
                if cell.value is not None:
                    val_str = str(cell.value).strip()
                    if val_str and val_str.lower() != 'nan':
                        try:
                            # Parse dengan parse_date_safe agar fleksibel
                            parsed = parse_date_safe(cell.value)
                            if pd.notna(parsed):
                                new_date = parsed.to_pydatetime().date()
                                # Cek jika ada perubahan nilai atau tipe/format
                                if cell.value != new_date or cell.number_format != 'yyyy-mm-dd':
                                    cell.value = new_date
                                    cell.number_format = 'yyyy-mm-dd'
                                    updated_count += 1
                                    print(f"✅ {full_name} - Tgl Lahir {label}: {val_str} -> {new_date.strftime('%Y-%m-%d')}")
                        except Exception as e:
                            print(f"⚠️ Gagal format tanggal untuk {full_name} ({label}): {val_str} - Error: {e}")
                            
    if updated_count > 0:
        wb.save(EXCEL_FILE)
        print(f"\n✅ Berhasil format {updated_count} tanggal lahir")
        tulis_log(f"Format tanggal lahir: {updated_count} tanggal berhasil diupdate")
    else:
        print("\n✅ Semua tanggal lahir sudah dalam format date yang benar")
        
    wb.close()
    print("="*60 + "\n")

def clean_address_string(address_str):
    if not address_str:
        return ""
    # 1. Ganti enter, carriage return, dan tab dengan spasi
    address_str = address_str.replace('\r', ' ').replace('\n', ' ').replace('\t', ' ')
    
    # 2. Hanya ijinkan karakter a-zA-Z0-9 .,/:\-()
    # Menghapus karakter di luar whitelist
    address_str = re.sub(r'[^a-zA-Z0-9 .,/:\-()]', '', address_str)
    
    # 3. Hilangkan spasi berlebih
    address_str = re.sub(r'\s+', ' ', address_str).strip()
    
    return address_str

def clean_addresses_in_excel():
    """Membersihkan format alamat di Excel (menghilangkan enter/newline dan karakter tidak valid)"""
    print("\n" + "="*60)
    print("🧹 MEMULAI PEMBERSIHAN ALAMAT DI EXCEL...")
    print("="*60)
    
    try:
        wb = load_workbook(EXCEL_FILE)
    except Exception as e:
        print(f"❌ Gagal membuka file Excel: {e}")
        print("="*60 + "\n")
        return
        
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    
    address_col_idx = headers.index('address') + 1 if 'address' in headers else None
    fullname_col_idx = headers.index('full_name') + 1 if 'full_name' in headers else None
    
    if not address_col_idx:
        print("❌ Kolom 'address' tidak ditemukan di Excel.")
        print("="*60 + "\n")
        wb.close()
        return
        
    updated_count = 0
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row_idx, address_col_idx)
        full_name = ws.cell(row_idx, fullname_col_idx).value if fullname_col_idx else 'N/A'
        
        if cell.value is not None:
            val_str = str(cell.value)
            cleaned_val = clean_address_string(val_str)
            
            if val_str != cleaned_val:
                cell.value = cleaned_val
                updated_count += 1
                visual_old = val_str.replace('\r', '\\r').replace('\n', '\\n')
                print(f"✅ {full_name} - Alamat dibersihkan: '{visual_old}' -> '{cleaned_val}'")
                
    if updated_count > 0:
        wb.save(EXCEL_FILE)
        print(f"\n✅ Berhasil membersihkan {updated_count} alamat.")
        tulis_log(f"Pembersihan alamat: {updated_count} alamat berhasil dibersihkan dari karakter ilegal/enter")
    else:
        print("\n✅ Semua alamat sudah dalam format yang valid.")
        
    wb.close()
    print("="*60 + "\n")
